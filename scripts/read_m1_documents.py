"""Utility script to extract useful text from the Phase M1 artefacts.

Outputs three helper files next to the originals:
- `M1 Full guide.extracted.txt`
- `Table.png.extracted.txt`
- `Copy_of_M1_...xlsx.extracted.txt`

Run with: python scripts/read_m1_documents.py
"""
from __future__ import annotations

import json
from pathlib import Path

from PyPDF2 import PdfReader
from openpyxl import load_workbook

try:
    import easyocr
except ImportError as exc:  # pragma: no cover - easyocr already installed via tooling
    raise SystemExit("easyocr is required. Install with 'pip install easyocr'") from exc

BASE_DIR = Path(__file__).resolve().parent.parent
DOCS_DIR = BASE_DIR / "M1 Complete guide and goal to achieve"

PDF_PATH = DOCS_DIR / "M1 Full guide.pdf"
PNG_PATH = DOCS_DIR / "Table.png"
XLSX_PATH = DOCS_DIR / "Copy_of_M1_ABHA_CREATION_AND_VERIFICATION_WITH_APIS_UPDATED_V1_2_7_Aug_1_58de4446bc.xlsx"


def extract_pdf() -> dict:
    if not PDF_PATH.exists():
        raise FileNotFoundError(f"Missing PDF at {PDF_PATH}")

    print(f"📄 Reading PDF: {PDF_PATH.name}")
    reader = PdfReader(PDF_PATH)
    output_path = PDF_PATH.with_suffix(".extracted.txt")

    highlights: list[str] = []
    with output_path.open("w", encoding="utf-8") as fh:
        for index, page in enumerate(reader.pages, start=1):
            text = page.extract_text() or ""
            fh.write(f"=== Page {index} ===\n{text}\n\n")

            for line in text.splitlines():
                stripped = line.strip()
                if not stripped:
                    continue
                # Capture key headings
                if any(keyword in stripped.lower() for keyword in ("objective", "goal", "feature", "phase", "workflow")):
                    highlights.append(stripped)

    print(f"✅ PDF text exported to {output_path}")
    return {
        "pages": len(reader.pages),
        "highlights": highlights[:50],  # limit to prevent huge console dump
        "output": output_path.name,
    }


def extract_png() -> dict:
    if not PNG_PATH.exists():
        raise FileNotFoundError(f"Missing PNG at {PNG_PATH}")

    print(f"🖼️ Reading PNG: {PNG_PATH.name}")
    reader = easyocr.Reader(["en"], gpu=False)
    results = reader.readtext(str(PNG_PATH), detail=0, paragraph=True)
    text = "\n".join(results)

    output_path = PNG_PATH.with_suffix(".extracted.txt")
    output_path.write_text(text, encoding="utf-8")
    print(f"✅ PNG text exported to {output_path}")

    return {
        "lines": len(results),
        "sample": results[:10],
        "output": output_path.name,
    }


def extract_xlsx() -> dict:
    if not XLSX_PATH.exists():
        raise FileNotFoundError(f"Missing XLSX at {XLSX_PATH}")

    print(f"📊 Reading XLSX: {XLSX_PATH.name}")
    workbook = load_workbook(XLSX_PATH, data_only=True, read_only=True)
    summary: dict[str, list[list[str]]] = {}

    for sheet in workbook.worksheets:
        rows: list[list[str]] = []
        for ridx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if ridx > 50:  # avoid massive dumps, capture first 50 rows
                break
            rows.append(["" if cell is None else str(cell).strip() for cell in row])
        summary[sheet.title] = rows

    output_path = XLSX_PATH.with_suffix(".extracted.json")
    output_path.write_text(json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"✅ XLSX preview exported to {output_path}")

    return {
        "sheets": list(summary.keys()),
        "output": output_path.name,
    }


def main() -> None:
    results = {
        "pdf": extract_pdf(),
        "png": extract_png(),
        "xlsx": extract_xlsx(),
    }

    summary_path = DOCS_DIR / "extraction_summary.json"
    summary_path.write_text(json.dumps(results, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"📝 Summary written to {summary_path}")


if __name__ == "__main__":
    main()
